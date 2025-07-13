#!/usr/bin/env python3
"""
lcsd_util_excel_timetable_parser.py
===================================
Standalone helper that downloads an LCSD *Field / Jogging Timetable* Excel
file and converts each relevant worksheet into a JSON–ready structure.

Key update (2025-07-05)
-----------------------
* When an Excel workbook contains **multiple viable timetable sheets**, the
  parser now detects each sheet’s month / year (from the title row or sheet
  name) and **only parses the sheet that matches the requested
  ``month_year``**.  
  If there is exactly one viable sheet, behaviour is unchanged.

Public API
~~~~~~~~~~
excel_to_timetable(excel_url: str,
                   month_year: str,
                   *,
                   sheet_keywords: tuple[str, ...] = (
                       "Field Timetable",
                       "Jogging Timetable"),
                   timeout: int = 15,
                   debug: bool = False) -> list[dict]

Return value – one dictionary per parsed worksheet:

    {
        "month_year":  "6/2025",
        "excel_url":   "<url>",
        "timetable":   {
            "YYYY-MM-DD": [
                {"start": "07:00", "end": "08:00", "status": "A"},
                …
            ],
            …
        },
        "legend_map":  { "A": "Available", … }
    }
"""
from __future__ import annotations

import datetime as _dt
import re as _re
from io import BytesIO as _BytesIO
from typing import Any, Dict, List, Tuple

import requests as _requests

try:
    from openpyxl import load_workbook as _load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover
    raise RuntimeError(
        "The 'openpyxl' library is required – install it via pip install openpyxl."
    ) from exc


###############################################################################
# ── constants & regexes ──
###############################################################################
_DEF_SHEET_KEYS: Tuple[str, ...] = ("Field Timetable", "Jogging Timetable")
_CODE_PATTERN = _re.compile(r"^[A-Z]$")  # legend code: single capital letter
_TIME_ROW_RE = _re.compile(r"^\s*\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}\s*$")
# new – detect “2025年7月”, “2025-07”, “07/2025”, etc.
_TITLE_DATE_RE = _re.compile(
    r"(?P<year>20\d{2})\s*(?:年|/|-)?\s*(?P<month>\d{1,2})\s*(?:月)?"
)


###############################################################################
# ── helpers ──
###############################################################################
def _download_excel(url: str, *, timeout: int, debug: bool) -> bytes:
    if debug:
        print(f"[DEBUG] Downloading Excel → {url}")
    resp = _requests.get(url, timeout=timeout)
    resp.raise_for_status()
    if debug:
        print(f"[DEBUG] Received {len(resp.content)} bytes.")
    return resp.content


def _detect_sheet_month_year(ws) -> str | None:
    """
    Try to detect the sheet’s «M/YYYY» string by scanning the first few rows and
    (as a fallback) the worksheet title.  Returns *None* if no pattern found.
    """
    # search first 5 rows
    for r in range(1, min(ws.max_row, 5) + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if isinstance(val, str):
                m = _TITLE_DATE_RE.search(val)
                if m:
                    year = int(m.group("year"))
                    month = int(m.group("month"))
                    return f"{month}/{year}"
    # fallback – sheet name
    m = _TITLE_DATE_RE.search(ws.title)
    if m:
        year = int(m.group("year"))
        month = int(m.group("month"))
        return f"{month}/{year}"
    return None


def _extract_legend(ws) -> Dict[str, str]:
    """Return mapping like {'A': 'Available', …} from rows above the '日期' row."""
    header_row = next(
        (r for r in range(1, ws.max_row + 1)
         if isinstance(ws.cell(r, 1).value, str) and "日期" in ws.cell(r, 1).value),
        1,
    )

    legend: Dict[str, str] = {}
    for r in range(1, header_row):
        code_col = None
        code_val = None
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if isinstance(val, str) and _CODE_PATTERN.fullmatch(val.strip()):
                code_col, code_val = c, val.strip()
                break
        if code_col:
            desc = " ".join(
                str(ws.cell(r, c).value).strip()
                for c in range(code_col + 1, ws.max_column + 1)
                if ws.cell(r, c).value not in (None, "")
            )
            if desc:
                legend[code_val] = desc
    return legend


def _parse_sheet(ws, month_year: str, *, debug: bool) -> Tuple[dict, dict]:
    """Parse a single worksheet into (timetable-dict, legend-dict)."""
    # 1️⃣ locate '日期' header row
    header_row = next(
        (
            r for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 1).value, str)
            and "日期" in ws.cell(r, 1).value
        ),
        1,
    )
    if debug:
        print(f"[DEBUG] '{ws.title}' header row → {header_row}")

    # 2️⃣ map columns → day-of-month
    month, year = map(int, month_year.split("/"))
    col_day: Dict[int, int] = {}
    for col in range(2, ws.max_column + 1):
        cell_val = ws.cell(header_row, col).value
        if isinstance(cell_val, (int, float)) and 1 <= int(cell_val) <= 31:
            col_day[col] = int(cell_val)
        elif isinstance(cell_val, str) and cell_val.strip().isdigit():
            d = int(cell_val.strip())
            if 1 <= d <= 31:
                col_day[col] = d
        else:  # fallback by position
            d = col - 1
            try:
                _dt.date(year, month, d)
            except ValueError:
                continue
            col_day[col] = d
    if debug:
        print(f"[DEBUG] column→day map: {col_day}")

    # 3️⃣ collect time-slot rows
    first_time_row = next(
        (
            r for r in range(header_row + 1, ws.max_row + 1)
            if isinstance(ws.cell(r, 1).value, str)
            and _TIME_ROW_RE.match(ws.cell(r, 1).value.strip())
        ),
        None,
    )
    if first_time_row is None:  # no timetable rows
        return {}, _extract_legend(ws)

    labels: List[str] = []
    rows: List[int] = []
    r = first_time_row
    while r <= ws.max_row and isinstance(ws.cell(r, 1).value, str) \
            and _TIME_ROW_RE.match(ws.cell(r, 1).value.strip()):
        labels.append(ws.cell(r, 1).value.split("-", 1)[0].strip())
        rows.append(r)
        r += 1
    first_label, last_label = labels[0], labels[-1]

    # 4️⃣ build per-date timetable
    timetable: Dict[str, List[dict]] = {}
    for col, day in col_day.items():
        try:
            date_iso = _dt.date(year, month, day).isoformat()
        except ValueError:
            continue

        raw: List[dict] = []
        for idx, row in enumerate(rows):
            start = labels[idx]
            end = labels[idx + 1] if idx + 1 < len(labels) else None
            if not end:
                continue
            cell_val = ws.cell(row, col).value
            status = str(cell_val).strip() if cell_val not in (None, "") else None
            if status:
                raw.append({"start": start, "end": end, "status": status})

        # merge consecutive identical statuses
        merged: List[dict] = []
        for itv in raw:
            if merged and merged[-1]["status"] == itv["status"] and merged[-1]["end"] == itv["start"]:
                merged[-1]["end"] = itv["end"]
            else:
                merged.append(itv.copy())

        # fill gaps with default "A"
        filled: List[dict] = []
        if merged:
            if merged[0]["start"] != first_label:
                filled.append({"start": first_label, "end": merged[0]["start"], "status": "A"})
            for i, itv in enumerate(merged):
                filled.append(itv)
                if i + 1 < len(merged) and itv["end"] != merged[i + 1]["start"]:
                    filled.append({"start": itv["end"], "end": merged[i + 1]["start"], "status": "A"})
            if merged[-1]["end"] != last_label:
                filled.append({"start": merged[-1]["end"], "end": last_label, "status": "A"})
        else:
            filled.append({"start": first_label, "end": last_label, "status": "A"})

        timetable[date_iso] = filled
        if debug:
            print(f"[DEBUG] {date_iso} – {len(filled)} intervals")

    return timetable, _extract_legend(ws)


###############################################################################
# ── public façade ──
###############################################################################
def excel_to_timetable(
    excel_url: str,
    month_year: str,
    *,
    sheet_keywords: Tuple[str, ...] = _DEF_SHEET_KEYS,
    timeout: int = 15,
    debug: bool = False,
) -> List[Dict[str, Any]]:
    """
    Download *excel_url*, parse every relevant worksheet, and return a list of
    timetable dictionaries (one per worksheet).

    Behaviour with multiple timetable sheets
    ----------------------------------------
    * If exactly **one** viable sheet exists → parsed as before.
    * If **multiple** viable sheets exist → only sheets whose month / year
      matches the supplied ``month_year`` are parsed.  If none match, the
      original behaviour (parse all viable sheets) is retained.
    """
    xls_bytes = _download_excel(excel_url, timeout=timeout, debug=debug)
    wb = _load_workbook(_BytesIO(xls_bytes), data_only=True)

    # pick candidate worksheets
    sheet_names = [n for n in wb.sheetnames if any(k in n for k in sheet_keywords)]
    if not sheet_names:
        sheet_names = [wb.active.title]  # fallback

    # drop '(New)' duplicates when a base exists
    cleaned: List[str] = []
    for name in sheet_names:
        base = name.replace("(New)", "").strip()
        if "(New)" in name and any(
            other for other in sheet_names
            if other.replace("(New)", "").strip() == base and other != name
        ):
            if debug:
                print(f"[DEBUG] Skipping '{name}' – '(New)' variant.")
            continue
        cleaned.append(name)
    if not cleaned:
        cleaned = sheet_names

    # ── NEW LOGIC: filter by month_year if multiple viable sheets ──
    if len(cleaned) > 1:
        tgt_month, tgt_year = map(int, month_year.split("/"))
        matched: List[str] = []
        for sheet_name in cleaned:
            ws = wb[sheet_name]
            detected = _detect_sheet_month_year(ws)
            if detected:
                dm, dy = map(int, detected.split("/"))
                if (dm, dy) == (tgt_month, tgt_year):
                    matched.append(sheet_name)
        if matched:
            if debug:
                print(f"[INFO] Multiple sheets found – keeping {matched} for {month_year}")
            sheet_names_to_parse = matched
        else:
            if debug:
                print("[WARN] No sheet matches requested month/year; parsing all viable sheets.")
            sheet_names_to_parse = cleaned
    else:
        sheet_names_to_parse = cleaned

    results: List[Dict[str, Any]] = []
    for sheet_name in sheet_names_to_parse:
        ws = wb[sheet_name]
        if debug:
            print(f"[INFO] Parsing worksheet '{sheet_name}'")
        timetable, legend = _parse_sheet(ws, month_year, debug=debug)
        results.append(
            {
                "month_year": month_year,
                "excel_url": excel_url,
                "timetable": timetable,
                "legend_map": legend,
            }
        )
    return results


__all__ = ["excel_to_timetable"]
