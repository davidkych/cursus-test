# ── src/routers/lcsd/timetable_endpoints.py ───────────────────────────
"""
LCSD athletic-field timetable builder – /api/lcsd/timetable

Behaviour
─────────
1. **Excel first, PDF fallback**  
   • For each facility the endpoint first tries to download and parse the
     monthly Excel sheet (legacy behaviour).  
   • If <u>no usable Excel data</u> are available for that facility it
     automatically falls back to the corresponding PDF, using the same
     parsing logic as *timetablepdf_endpoints.py*.

2. **Graceful degradation & diagnostics** (unchanged)  
   • All errors are collected in a `diagnostics` list; robust facilities
     are still returned.  
   • If *nothing* can be produced for <em>any</em> facility the endpoint
     responds with HTTP 500 and a descriptive error.

3. **JSON format**  
   • Identical to the previous Excel version:  
       { metadata, source, facilities:{…} }  
     Each facility’s `schedules` array contains either Excel-based items  
     *(fields: sheet_name, excel_url, …)* or PDF-based items  
     *(fields: pdf_url, …)* depending on the source actually used.
"""

from fastapi import APIRouter, Query
from fastapi.responses import JSONResponse
from pydantic import BaseModel, Field
from typing import Dict, List, Any, Tuple, Optional
import datetime, hashlib, logging, os, re, requests
from io import BytesIO

from azure.cosmos import CosmosClient
from azure.identity import DefaultAzureCredential
from openpyxl import load_workbook
import pdfplumber                                    # ← NEW

# ── Cosmos helpers ────────────────────────────────────────────────────
TAG_PROBE     = "timetableprobe"
TAG_TIMETABLE = "timetable"

_cosmos_endpoint = os.getenv("COSMOS_ENDPOINT")
_database_name   = os.getenv("COSMOS_DATABASE", "cursusdb")
_container_name  = os.getenv("COSMOS_CONTAINER", "jsonContainer")
_cosmos_key      = os.getenv("COSMOS_KEY")

_client = (
    CosmosClient(_cosmos_endpoint, credential=_cosmos_key)
    if _cosmos_key else CosmosClient(_cosmos_endpoint, credential=DefaultAzureCredential())
)
_container = _client.get_database_client(_database_name).get_container_client(_container_name)

def _item_id(tag, secondary, year, month, day=None) -> str:
    parts = [tag, secondary, str(year), str(month)]
    if day is not None:
        parts.append(str(day))
    return "_".join(parts)

def _upsert(tag, secondary, year, month, day, data):
    _container.upsert_item({
        "id": _item_id(tag, secondary, year, month, day),
        "tag": tag,
        "secondary_tag": secondary,
        "year": year,
        "month": month,
        "day": day,
        "data": data,
    })

# ── Excel-parsing utilities (unchanged) ──────────────────────────────
log = logging.getLogger("lcsd.timetable")
log.setLevel(logging.INFO)

_CODE_RE = re.compile(r"^[A-Z]$")
_TIME_RE = re.compile(r"^\s*\d{1,2}:\d{2}\s*[–-]\s*\d{1,2}:\d{2}\s*$")  # hyphen or en-dash

def _extract_legend(ws) -> Dict[str, str]:
    header_row = next(
        (
            r
            for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 1).value, str) and "日期" in ws.cell(r, 1).value
        ),
        1,
    )
    legend = {}
    for r in range(1, header_row):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and _CODE_RE.match(v.strip()):
                desc = " ".join(
                    str(ws.cell(r, cc).value).strip()
                    for cc in range(c + 1, ws.max_column + 1)
                    if ws.cell(r, cc).value not in (None, "")
                )
                legend[v.strip()] = desc
                break
    return legend

def _parse_sheet(ws, month: int, year: int) -> Tuple[Dict[str, Any], Dict[str, str]]:
    header_row = next(
        (
            r
            for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 1).value, str) and "日期" in ws.cell(r, 1).value
        ),
        1,
    )

    # map columns → day
    col2day: Dict[int, int] = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, (int, float)) and 1 <= int(v) <= 31:
            col2day[c] = int(v)
        elif isinstance(v, str) and v.strip().isdigit():
            d = int(v.strip())
            if 1 <= d <= 31:
                col2day[c] = d
        else:
            # infer from column index
            inferred = c - 1
            try:
                datetime.date(year, month, inferred)
                col2day[c] = inferred
            except ValueError:
                pass

    # consecutive time rows
    time_rows, time_labels = [], []
    for r in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(r, 1).value
        if isinstance(cell, str) and _TIME_RE.match(cell.strip()):
            time_rows.append(r)
            time_labels.append(cell.split("–" if "–" in cell else "-", 1)[0].strip())
        elif time_rows:
            break
    if not time_rows:
        return {}, _extract_legend(ws)

    timetable: Dict[str, List[Dict[str, str]]] = {}
    for c, day in col2day.items():
        try:
            date_iso = datetime.date(year, month, day).isoformat()
        except ValueError:
            continue

        raw: List[Dict[str, str]] = []
        for idx, r in enumerate(time_rows):
            start = time_labels[idx]
            end   = time_labels[idx + 1] if idx + 1 < len(time_rows) else None
            if not end:
                continue
            status = ws.cell(r, c).value
            status = str(status).strip() if status else ""
            if status:
                raw.append({"start": start, "end": end, "status": status})

        # merge contiguous identical status blocks
        merged: List[Dict[str, str]] = []
        for itv in raw:
            if merged and merged[-1]["status"] == itv["status"] and merged[-1]["end"] == itv["start"]:
                merged[-1]["end"] = itv["end"]
            else:
                merged.append(itv.copy())

        # fill gaps with status "A"
        filled: List[Dict[str, str]] = []
        if merged:
            if merged[0]["start"] != time_labels[0]:
                filled.append({"start": time_labels[0], "end": merged[0]["start"], "status": "A"})
            for i, itv in enumerate(merged):
                filled.append(itv)
                nxt = merged[i + 1]["start"] if i + 1 < len(merged) else None
                if nxt and itv["end"] != nxt:
                    filled.append({"start": itv["end"], "end": nxt, "status": "A"})
            if merged[-1]["end"] != time_labels[-1]:
                filled.append({"start": merged[-1]["end"], "end": time_labels[-1], "status": "A"})
        else:
            filled.append({"start": time_labels[0], "end": time_labels[-1], "status": "A"})

        timetable[date_iso] = filled

    return timetable, _extract_legend(ws)

# ── PDF-parsing utilities (ported from timetablepdf_endpoints) ────────
_TIME_ROW_RE = re.compile(r"^\s*(\d{1,2}:\d{2})\s*[–-]\s*(\d{1,2}:\d{2})\s+(.*)$")
_LEGEND_RE   = re.compile(r"^\s*([A-Z])\s+(.+)$")

def _parse_pdf_bytes(pdf_bytes: bytes, month: int, year: int) -> Tuple[Dict[str, Any], Dict[str, str]]:
    lines: List[str] = []
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for p in pdf.pages:
            lines.extend((p.extract_text() or "").splitlines())

    first_time_idx: Optional[int] = None
    legend_map: Dict[str, str] = {}
    for idx, ln in enumerate(lines):
        if _TIME_ROW_RE.match(ln):
            first_time_idx = idx
            break
        m = _LEGEND_RE.match(ln)
        if m and m.group(1) not in legend_map:
            legend_map[m.group(1)] = m.group(2).strip()

    if first_time_idx is None:
        return {}, legend_map

    time_labels: List[str] = []
    status_matrix: List[List[str]] = []
    for ln in lines[first_time_idx:]:
        m = _TIME_ROW_RE.match(ln)
        if not m:
            if time_labels:
                break
            continue
        start_lbl, _end_lbl, rest = m.groups()
        tokens = rest.strip().split()
        time_labels.append(start_lbl)
        status_matrix.append(tokens)

    if not status_matrix:
        return {}, legend_map

    col_count = max(len(r) for r in status_matrix)
    for r in status_matrix:
        r += [""] * (col_count - len(r))

    timetable: Dict[str, List[Dict[str, str]]] = {}
    for col in range(col_count):
        day = col + 1
        try:
            date_iso = datetime.date(year, month, day).isoformat()
        except ValueError:
            continue

        raw: List[Dict[str, str]] = []
        for i, row in enumerate(status_matrix):
            status = row[col].strip()
            if not status:
                continue
            start = time_labels[i]
            end   = time_labels[i + 1] if i + 1 < len(time_labels) else None
            if end:
                raw.append({"start": start, "end": end, "status": status})

        merged: List[Dict[str, str]] = []
        for iv in raw:
            if merged and merged[-1]["status"] == iv["status"] and merged[-1]["end"] == iv["start"]:
                merged[-1]["end"] = iv["end"]
            else:
                merged.append(iv.copy())

        filled: List[Dict[str, str]] = []
        if merged:
            if merged[0]["start"] != time_labels[0]:
                filled.append({"start": time_labels[0], "end": merged[0]["start"], "status": "A"})
            for i, iv in enumerate(merged):
                filled.append(iv)
                nxt = merged[i + 1]["start"] if i + 1 < len(merged) else None
                if nxt and iv["end"] != nxt:
                    filled.append({"start": iv["end"], "end": nxt, "status": "A"})
            if merged[-1]["end"] != time_labels[-1]:
                filled.append({"start": merged[-1]["end"], "end": time_labels[-1], "status": "A"})
        else:
            filled.append({"start": time_labels[0], "end": time_labels[-1], "status": "A"})

        timetable[date_iso] = filled

    return timetable, legend_map

# ── Misc helpers ─────────────────────────────────────────────────────
def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()

def _latest_probe() -> Dict[str, Any]:
    docs = list(
        _container.query_items(
            query="""
                SELECT TOP 1 c.data
                FROM   c
                WHERE  c.tag = 'lcsd' AND c.secondary_tag = @s
                ORDER BY c._ts DESC
            """,
            parameters=[{"name": "@s", "value": TAG_PROBE}],
            enable_cross_partition_query=True,
        )
    )
    if not docs:
        raise RuntimeError("No timetable-probe data found in Cosmos DB.")
    return docs[0]["data"]

# ── Core builder ─────────────────────────────────────────────────────
def _build_timetable(month: int, year: int) -> Tuple[Dict[str, Any], List[str]]:
    probe = _latest_probe()
    facilities_out: Dict[str, Any] = {}
    diagnostics: List[str] = []

    for facility in probe.get("timetables", []):
        lcsd_no   = facility.get("lcsd_number")
        base_name = facility.get("name")
        schedules = facility.get("jogging_schedule", [])

        # -- 1  Try Excel --------------------------------------------------
        excel_results: List[Dict[str, Any]] = []
        for sch in schedules:
            try:
                m = int(str(sch.get("month")).lstrip("0") or "0")
                y = int(sch.get("year"))
            except Exception:
                diagnostics.append(f"{lcsd_no}: invalid month/year in probe")
                continue
            if m != month or y != year:
                continue

            xurl = sch.get("excel_url")
            if not xurl:
                continue

            try:
                resp = requests.get(xurl, timeout=15)
                resp.raise_for_status()
            except requests.RequestException as exc:
                diagnostics.append(f"{lcsd_no}: Excel download failed – {exc}")
                continue

            try:
                wb = load_workbook(BytesIO(resp.content), data_only=True)
            except Exception as exc:
                diagnostics.append(f"{lcsd_no}: openpyxl load failed – {exc}")
                continue

            cand = (
                [n for n in wb.sheetnames if "Field Timetable" in n]
                or [n for n in wb.sheetnames if "Jogging Timetable" in n]
                or [wb.active.title]
            )
            sheets = [
                n
                for n in cand
                if not ("(New)" in n and n.replace("(New)", "").strip() in cand)
            ]

            for sname in sheets:
                try:
                    tt, legend = _parse_sheet(wb[sname], month, year)
                except Exception as exc:
                    diagnostics.append(f"{lcsd_no}/{sname}: Excel parse error – {exc}")
                    continue

                excel_results.append(
                    {
                        "sheet_name": sname,
                        "excel_url": xurl,
                        "sha256": _sha256(resp.content),
                        "timetable": tt,
                        "legend_map": legend,
                    }
                )

        # -- 2  Fallback to PDF if Excel failed ----------------------------
        pdf_results: List[Dict[str, Any]] = []
        if not excel_results:
            for sch in schedules:
                try:
                    m = int(str(sch.get("month")).lstrip("0") or "0")
                    y = int(sch.get("year"))
                except Exception:
                    continue
                if m != month or y != year:
                    continue

                source = sch.get("pdf_url") or sch.get("pdf_filename") or ""
                if not source:
                    continue

                try:
                    if source.startswith("http"):
                        resp = requests.get(source, timeout=15)
                        resp.raise_for_status()
                        pdf_bytes = resp.content
                    else:
                        with open(source, "rb") as fh:
                            pdf_bytes = fh.read()
                except Exception as exc:
                    diagnostics.append(f"{lcsd_no}: PDF fetch failed – {exc}")
                    continue

                try:
                    tt, legend = _parse_pdf_bytes(pdf_bytes, month, year)
                except Exception as exc:
                    diagnostics.append(f"{lcsd_no}: PDF parse error – {exc}")
                    continue

                pdf_results.append(
                    {
                        "pdf_url": source,
                        "sha256": _sha256(pdf_bytes),
                        "timetable": tt,
                        "legend_map": legend,
                    }
                )

        # -- 3  Decide what to store ---------------------------------------
        if excel_results:
            sch_results = excel_results
        elif pdf_results:
            sch_results = pdf_results
        else:
            diagnostics.append(f"{lcsd_no}: no usable sheet/PDF for {month:02}/{year}")
            continue

        if len(sch_results) == 1:
            facilities_out[lcsd_no] = {"name": base_name, "schedules": sch_results}
        else:
            for idx, sr in enumerate(sch_results):
                key = f"{lcsd_no}{chr(ord('a') + idx)}"
                facilities_out[key] = {
                    "name": base_name,
                    "schedules": [sr],
                }

    if not facilities_out:
        raise RuntimeError(
            f"No facility produced a timetable for {month:02}/{year} – see diagnostics"
        )

    ts = datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z"
    data = {
        "metadata": {"timestamp": ts, "year": year, "month": month},
        "source": TAG_PROBE,
        "facilities": facilities_out,
    }
    return data, diagnostics

def _store_and_prepare_response(data: Dict[str, Any], diags: List[str], year: int, month: int):
    _upsert("lcsd", TAG_TIMETABLE, year, month, None, data)
    return {
        "stored_id":      _item_id("lcsd", TAG_TIMETABLE, year, month),
        "facility_count": len(data["facilities"]),
        "diagnostics":    diags,
    }

# ── FastAPI router ───────────────────────────────────────────────────
class TimetableReq(BaseModel):
    year: int  = Field(..., ge=1900)
    month: int = Field(..., ge=1, le=12)

router = APIRouter()

@router.post("/api/lcsd/timetable", summary="Build LCSD timetable JSON (POST)")
def timetable_post(req: TimetableReq):
    try:
        data, diags = _build_timetable(req.month, req.year)
        return _store_and_prepare_response(data, diags, req.year, req.month)
    except Exception as exc:
        return JSONResponse(status_code=500, content={"status": "error", "detail": str(exc)})

@router.get("/api/lcsd/timetable", summary="Build LCSD timetable JSON (GET)")
def timetable_get(
    year: int = Query(..., ge=1900),
    month: int = Query(..., ge=1, le=12),
):
    try:
        data, diags = _build_timetable(month, year)
        return _store_and_prepare_response(data, diags, year, month)
    except Exception as exc:
        return JSONResponse(status_code=500, content={"status": "error", "detail": str(exc)})

# ── End of file ──────────────────────────────────────────────────────
